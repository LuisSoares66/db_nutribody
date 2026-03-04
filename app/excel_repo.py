import os
from openpyxl import load_workbook
from datetime import date, datetime
import sys

def _base_dir():
    # quando for exe (PyInstaller), sys._MEIPASS existe
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)  # pasta do exe
    # modo normal (python)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

BASE_DIR = _base_dir()
DATA_FILE = os.path.join(BASE_DIR, "data", "backup_nutri_hospital.xlsx")

SHEET_H = "Hospitais"
SHEET_C = "Contatos"
SHEET_D = "DadosHospital"
SHEET_P = "ProdutosHospital"

SHEET_V = "Visitas"

VISITAS_COLS = [
    "id",
    "hospital_id",
    "nome_hospital",
    "data_visita",
    "data_retorno",
    "observacao",
    "criado_em",
]

def _ensure_hospitais_headers(ws):
    """
    Garante que a aba Hospitais tenha as colunas:
    ... estado, data_visita, data_retorno (nessa ordem).
    Se não existir, cria e posiciona logo após 'estado'.
    """
    headers, _ = _sheet(ws)
    if not headers:
        raise ValueError("A aba 'Hospitais' precisa ter cabeçalho na linha 1.")

    def add_after(col_name, new_name):
        nonlocal headers
        if new_name in headers:
            return
        if col_name not in headers:
            # se não achar 'estado', adiciona no final
            headers.append(new_name)
            ws.cell(row=1, column=len(headers)).value = new_name
            return

        idx = headers.index(col_name) + 1  # posição após col_name (0-based -> insert)
        headers.insert(idx, new_name)

        # reescreve o cabeçalho inteiro (linha 1) na ordem nova
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c).value = h

    # garante em ordem após estado
    add_after("estado", "data_visita")
    # agora data_retorno deve vir depois de data_visita
    add_after("data_visita", "data_retorno")

    return headers

def _wb():
    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError(f"Não achei o Excel: {DATA_FILE}")
    return load_workbook(DATA_FILE)

def _sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]
    return headers, data

def _next_id(values):
    mx = 0
    for v in values:
        try:
            mx = max(mx, int(v or 0))
        except:
            pass
    return mx + 1

def _save(wb):
    wb.save(DATA_FILE)


# =========================
# HOSPITAIS
# =========================
def list_hospitais(order_by="nome"):
    wb = _wb()
    ws = wb[SHEET_H]

    _ensure_hospitais_headers(ws)
    _save(wb)  # ✅ garante que as colunas fiquem gravadas no arquivo
    headers, data = _sheet(ws)
    items = [dict(zip(headers, r)) for r in data if any(r)]

    key = "nome_hospital" if order_by == "nome" else "cidade"
    items.sort(key=lambda x: str(x.get(key) or "").strip().lower())
    return items

def get_hospital(hospital_id: int):
    wb = _wb()
    ws = wb[SHEET_H]

    _ensure_hospitais_headers(ws)
    _save(wb)  # ✅
    headers, data = _sheet(ws)

    for r in data:
        if not r or r[0] is None:
            continue
        if int(r[0]) == int(hospital_id):
            return dict(zip(headers, r))
    return None

def save_hospital(payload: dict):
    wb = _wb()
    ws = wb[SHEET_H]

    headers = _ensure_hospitais_headers(ws)
    _save(wb)  # ✅ garante que o cabeçalho atualizado fique persistido

    if not headers or "id" not in headers:
        raise ValueError("A aba 'Hospitais' precisa ter cabeçalho com coluna 'id'.")

    id_col = headers.index("id") + 1
    wanted_id = payload.get("id")
    if wanted_id:
        wanted_id = int(wanted_id)

    row_to_update = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        if v is not None and int(v) == wanted_id:
            row_to_update = r
            break

    if row_to_update is None:
        ids = [ws.cell(r, id_col).value for r in range(2, ws.max_row + 1)]
        new_id = _next_id(ids)
        payload["id"] = new_id
        row_to_update = ws.max_row + 1

    # ✅ garante que as chaves existam no payload (se não vier, salva vazio)
    payload.setdefault("data_visita", "")
    payload.setdefault("data_retorno", "")

    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=row_to_update, column=col_i).value = payload.get(h)

    _save(wb)
    return int(payload["id"])

def delete_hospital(hospital_id: int):
    wb = _wb()

    # Hospitais
    ws = wb[SHEET_H]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == hospital_id:
            ws.delete_rows(r)

    # Contatos (hospital_id = coluna 2)
    ws = wb[SHEET_C]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 2).value == hospital_id:
            ws.delete_rows(r)

    # DadosHospital (hospital_id = coluna 2)
    ws = wb[SHEET_D]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 2).value == hospital_id:
            ws.delete_rows(r)

    # ProdutosHospital (hospital_id = coluna 2)
    ws = wb[SHEET_P]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 2).value == hospital_id:
            ws.delete_rows(r)
    # Visitas (hospital_id = coluna 2)
    if SHEET_V in wb.sheetnames:
        ws = wb[SHEET_V]
        for r in range(ws.max_row, 1, -1):
            if ws.cell(r, 2).value == hospital_id:
                ws.delete_rows(r)
    _save(wb)


# =========================
# CONTATOS
# =========================
def list_contatos(hospital_id: int):
    wb = _wb()
    ws = wb[SHEET_C]
    headers, data = _sheet(ws)
    out = []
    for r in data:
        if not r or r[0] is None:
            continue
        if int(r[1] or 0) == int(hospital_id):
            out.append(dict(zip(headers, r)))
    return out

def save_contato(payload: dict):
    wb = _wb()
    ws = wb[SHEET_C]
    headers, _ = _sheet(ws)
    if not headers or "id" not in headers:
        raise ValueError("A aba 'Contatos' precisa ter coluna 'id'.")

    id_col = headers.index("id") + 1
    wanted_id = payload.get("id")
    if wanted_id:
        wanted_id = int(wanted_id)

    row_to_update = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        if v is not None and int(v) == wanted_id:
            row_to_update = r
            break

    if row_to_update is None:
        ids = [ws.cell(r, id_col).value for r in range(2, ws.max_row + 1)]
        new_id = _next_id(ids)
        payload["id"] = new_id
        row_to_update = ws.max_row + 1

    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=row_to_update, column=col_i).value = payload.get(h)

    _save(wb)
    return int(payload["id"])

def delete_contato(contato_id: int):
    wb = _wb()
    ws = wb[SHEET_C]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == contato_id:
            ws.delete_rows(r)
    _save(wb)


# =========================
# DADOS HOSPITAL
# =========================
def get_dados(hospital_id: int):
    wb = _wb()
    ws = wb[SHEET_D]
    headers, data = _sheet(ws)
    for r in data:
        if not r or r[0] is None:
            continue
        if int(r[1] or 0) == int(hospital_id):
            return dict(zip(headers, r))
    return None

def save_dados(payload: dict):
    wb = _wb()
    ws = wb[SHEET_D]
    headers, _ = _sheet(ws)
    if not headers or "hospital_id" not in headers:
        raise ValueError("A aba 'DadosHospital' precisa ter coluna 'hospital_id'.")

    hid_col = headers.index("hospital_id") + 1
    row_to_update = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, hid_col).value
        if v is not None and int(v) == int(payload["hospital_id"]):
            row_to_update = r
            break

    if row_to_update is None:
        ws.append([payload.get(h) for h in headers])
    else:
        for col_i, h in enumerate(headers, start=1):
            ws.cell(row=row_to_update, column=col_i).value = payload.get(h)

    _save(wb)


# =========================
# PRODUTOS
# =========================
def list_produtos(hospital_id: int):
    wb = _wb()
    ws = wb[SHEET_P]
    headers, data = _sheet(ws)
    out = []
    for r in data:
        if not r or r[0] is None:
            continue
        if int(r[1] or 0) == int(hospital_id):
            out.append(dict(zip(headers, r)))
    return out

def save_produto(payload: dict):
    wb = _wb()
    ws = wb[SHEET_P]
    headers, _ = _sheet(ws)
    if not headers or "id" not in headers:
        raise ValueError("A aba 'ProdutosHospital' precisa ter coluna 'id'.")

    id_col = headers.index("id") + 1
    wanted_id = payload.get("id")
    if wanted_id:
        wanted_id = int(wanted_id)

    row_to_update = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        if v is not None and int(v) == wanted_id:
            row_to_update = r
            break

    if row_to_update is None:
        ids = [ws.cell(r, id_col).value for r in range(2, ws.max_row + 1)]
        new_id = _next_id(ids)
        payload["id"] = new_id
        row_to_update = ws.max_row + 1

    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=row_to_update, column=col_i).value = payload.get(h)

    _save(wb)
    return int(payload["id"])

def delete_produto(produto_id: int):
    wb = _wb()
    ws = wb[SHEET_P]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == produto_id:
            ws.delete_rows(r)
    _save(wb)
    
def _ensure_visitas_sheet(wb):
    if SHEET_V not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_V)
        for i, h in enumerate(VISITAS_COLS, start=1):
            ws.cell(row=1, column=i).value = h
        return ws

    ws = wb[SHEET_V]
    headers, _ = _sheet(ws)
    if not headers:
        for i, h in enumerate(VISITAS_COLS, start=1):
            ws.cell(row=1, column=i).value = h
    else:
        # garante colunas que faltam e reescreve header (ordem fixa)
        for h in VISITAS_COLS:
            if h not in headers:
                headers.append(h)
        for i, h in enumerate(VISITAS_COLS, start=1):
            ws.cell(row=1, column=i).value = h
    return ws

def _to_int(v, default=0):
    try:
        if v is None:
            return default
        s = str(v).strip()
        if s == "":
            return default
        return int(float(s))
    except:
        return default

def list_visitas(hospital_id: int):
    wb = _wb()
    ws = _ensure_visitas_sheet(wb)
    headers, data = _sheet(ws)

    items = []
    for r in data:
        if not r or not any(r):
            continue
        d = dict(zip(headers, r))
        if _to_int(d.get("hospital_id"), 0) == _to_int(hospital_id, 0):
            items.append(d)

    # ordena por data_retorno/data_visita
    items.sort(key=lambda x: (str(x.get("data_retorno") or ""), str(x.get("data_visita") or "")))
    return items

def save_visita(payload: dict):
    wb = _wb()
    ws = _ensure_visitas_sheet(wb)
    headers, _ = _sheet(ws)

    id_col = headers.index("id") + 1

    wanted_id = payload.get("id")
    wanted_id = int(wanted_id) if wanted_id else None

    row_to_update = None
    if wanted_id:
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, id_col).value
            if v is not None and int(v) == wanted_id:
                row_to_update = r
                break

    if row_to_update is None:
        ids = [ws.cell(r, id_col).value for r in range(2, ws.max_row + 1)]
        new_id = _next_id(ids)
        payload["id"] = new_id
        row_to_update = ws.max_row + 1

    payload.setdefault("observacao", "")
    payload.setdefault("criado_em", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=row_to_update, column=col_i).value = payload.get(h, "")

    _save(wb)
    return int(payload["id"])

def delete_visita(visita_id: int):
    wb = _wb()
    ws = _ensure_visitas_sheet(wb)
    headers, _ = _sheet(ws)
    id_col = headers.index("id") + 1

    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        if v is not None and int(v) == int(visita_id):
            ws.delete_rows(r, 1)
            _save(wb)
            return True
    return False
